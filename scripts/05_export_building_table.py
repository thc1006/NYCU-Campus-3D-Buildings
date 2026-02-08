"""
Export NYCU building data to CSV and Excel with computed dimensions.

Reads NYCU_buildings_3d.geojson and computes:
- Bounding box width/length (meters)
- Footprint area (shoelace formula, m^2)
- Estimated total floor area (footprint * floors)

Outputs:
- NYCU_buildings_table.csv
- NYCU_buildings_table.xlsx
"""
import csv
import json
import math
import os
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

# Conversion factors at latitude ~24.787
DEG_LON_TO_M = 101235.0  # 1 degree longitude in meters
DEG_LAT_TO_M = 110574.0  # 1 degree latitude in meters


def shoelace_area_m2(coords):
    """Compute polygon area in m^2 using shoelace formula with lat/lon."""
    n = len(coords)
    if n < 3:
        return 0.0
    # Close polygon if needed
    if coords[0] != coords[-1]:
        coords = coords + [coords[0]]
        n += 1

    area = 0.0
    for i in range(n - 1):
        x1 = coords[i][0] * DEG_LON_TO_M
        y1 = coords[i][1] * DEG_LAT_TO_M
        x2 = coords[i + 1][0] * DEG_LON_TO_M
        y2 = coords[i + 1][1] * DEG_LAT_TO_M
        area += x1 * y2 - x2 * y1
    return abs(area) / 2.0


def bounding_box_dims(coords):
    """Compute bounding box width and length in meters."""
    if not coords:
        return 0.0, 0.0
    lons = [c[0] for c in coords]
    lats = [c[1] for c in coords]
    width = (max(lons) - min(lons)) * DEG_LON_TO_M
    length = (max(lats) - min(lats)) * DEG_LAT_TO_M
    return round(width, 1), round(length, 1)


def main():
    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(project_dir, 'data')

    # Read 3D GeoJSON
    geojson_file = os.path.join(data_dir, 'output', 'NYCU_buildings_3d.geojson')
    with open(geojson_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    features = data['features']
    print(f'Loaded {len(features)} buildings from {geojson_file}')

    # Build rows
    rows = []
    for feat in features:
        props = feat['properties']
        geom = feat['geometry']

        if geom['type'] != 'Polygon':
            continue

        ring = geom['coordinates'][0]
        name = props.get('display_name', '')
        name_en = props.get('display_name_en', '')
        height = props.get('height_m', 0)
        floors_raw = props.get('building:levels', '')
        nlsc_h = props.get('nlsc_BUILD_H', '')
        build_str = props.get('nlsc_BUILD_STR', '')
        build_id = props.get('nlsc_BUILD_ID', '')
        building_type = props.get('building', '')
        osm_id = props.get('osm_id', '')

        # Parse floors
        floors = ''
        floors_num = 0
        if floors_raw:
            floors = str(floors_raw)
            try:
                floors_num = int(str(floors_raw).replace('>', '').strip())
            except ValueError:
                floors_num = 0

        # Compute dimensions
        width, length = bounding_box_dims(ring)
        area = round(shoelace_area_m2(ring), 1)

        # Estimate total floor area
        total_floor_area = ''
        if floors_num > 0 and area > 0:
            total_floor_area = round(area * floors_num, 0)

        # Structure type mapping
        str_map = {
            'R': '鋼筋混凝土(RC)',
            'S': '鋼骨(S)',
            'B': '磚造',
            'W': '木造',
            'M': '金屬/混合',
            'T': '臨時',
        }
        structure = str_map.get(build_str, build_str) if build_str else ''

        rows.append({
            'name': name,
            'name_en': name_en,
            'height_m': height,
            'floors': floors,
            'width_m': width,
            'length_m': length,
            'footprint_area_m2': area,
            'total_floor_area_m2': total_floor_area,
            'structure': structure,
            'building_type': building_type,
            'nlsc_BUILD_H': nlsc_h,
            'nlsc_BUILD_ID': build_id,
            'osm_id': osm_id,
        })

    # Sort by height descending
    rows.sort(key=lambda r: -float(r['height_m'] or 0))

    print(f'Processed {len(rows)} polygon buildings')

    # --- CSV output ---
    csv_file = os.path.join(data_dir, 'output', 'NYCU_buildings_table.csv')
    fieldnames = [
        'name', 'name_en', 'height_m', 'floors',
        'width_m', 'length_m', 'footprint_area_m2', 'total_floor_area_m2',
        'structure', 'building_type', 'nlsc_BUILD_H', 'nlsc_BUILD_ID', 'osm_id',
    ]
    headers_zh = {
        'name': '建築名稱',
        'name_en': '英文名稱',
        'height_m': '高度(m)',
        'floors': '樓層',
        'width_m': '東西寬(m)',
        'length_m': '南北長(m)',
        'footprint_area_m2': '投影面積(m²)',
        'total_floor_area_m2': '估算總樓地板面積(m²)',
        'structure': '結構類型',
        'building_type': 'OSM建築類型',
        'nlsc_BUILD_H': 'NLSC高度',
        'nlsc_BUILD_ID': 'NLSC建物ID',
        'osm_id': 'OSM ID',
    }

    with open(csv_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        # Write Chinese headers
        writer.writerow(headers_zh)
        writer.writerows(rows)
    print(f'Saved CSV: {csv_file}')

    # --- Excel output ---
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Sheet 1: All buildings
        ws = wb.active
        ws.title = '所有建築'

        # Header style
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Write headers
        zh_headers = [headers_zh[f] for f in fieldnames]
        for col_idx, header in enumerate(zh_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                val = row_data[field]
                # Convert numeric strings to numbers for Excel
                if field in ('height_m', 'width_m', 'length_m', 'footprint_area_m2'):
                    try:
                        val = float(val) if val else ''
                    except (ValueError, TypeError):
                        pass
                elif field == 'total_floor_area_m2':
                    try:
                        val = int(float(val)) if val else ''
                    except (ValueError, TypeError):
                        pass

                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if field in ('height_m', 'width_m', 'length_m', 'footprint_area_m2',
                             'total_floor_area_m2', 'floors'):
                    cell.alignment = Alignment(horizontal='right')

        # Auto-adjust column widths
        col_widths = {
            'name': 22, 'name_en': 35, 'height_m': 10, 'floors': 8,
            'width_m': 12, 'length_m': 12, 'footprint_area_m2': 16,
            'total_floor_area_m2': 20, 'structure': 16, 'building_type': 14,
            'nlsc_BUILD_H': 12, 'nlsc_BUILD_ID': 14, 'osm_id': 12,
        }
        for col_idx, field in enumerate(fieldnames, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 12)

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f'A1:{get_column_letter(len(fieldnames))}{len(rows) + 1}'

        # Sheet 2: Named buildings only
        ws2 = wb.create_sheet('有名稱建築')
        named_rows = [r for r in rows if r['name']]

        for col_idx, header in enumerate(zh_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_idx, row_data in enumerate(named_rows, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                val = row_data[field]
                if field in ('height_m', 'width_m', 'length_m', 'footprint_area_m2'):
                    try:
                        val = float(val) if val else ''
                    except (ValueError, TypeError):
                        pass
                elif field == 'total_floor_area_m2':
                    try:
                        val = int(float(val)) if val else ''
                    except (ValueError, TypeError):
                        pass
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        for col_idx, field in enumerate(fieldnames, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 12)
        ws2.freeze_panes = 'A2'
        ws2.auto_filter.ref = f'A1:{get_column_letter(len(fieldnames))}{len(named_rows) + 1}'

        # Sheet 3: Summary statistics
        ws3 = wb.create_sheet('統計摘要')
        stats = [
            ('項目', '數值'),
            ('建築總數（多邊形）', len(rows)),
            ('有名稱建築', len(named_rows)),
            ('有NLSC高度資料', sum(1 for r in rows if r['nlsc_BUILD_H'])),
            ('有樓層資料', sum(1 for r in rows if r['floors'])),
            ('', ''),
            ('最高建築', f"{rows[0]['name']} ({rows[0]['height_m']}m)" if rows else ''),
            ('最大投影面積', ''),
            ('平均高度(m)', round(sum(float(r['height_m']) for r in rows if r['height_m']) / max(1, len([r for r in rows if r['height_m']])), 1)),
            ('平均投影面積(m²)', round(sum(float(r['footprint_area_m2']) for r in rows if r['footprint_area_m2']) / max(1, len([r for r in rows if r['footprint_area_m2']])), 1)),
        ]

        # Find max area building
        max_area_row = max(rows, key=lambda r: float(r['footprint_area_m2'] or 0))
        stats[7] = ('最大投影面積', f"{max_area_row['name']} ({max_area_row['footprint_area_m2']}m²)")

        for row_idx, (label, value) in enumerate(stats, 1):
            ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True) if label else Font()
            ws3.cell(row=row_idx, column=2, value=value)
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 40

        xlsx_file = os.path.join(data_dir, 'output', 'NYCU_buildings_table.xlsx')
        wb.save(xlsx_file)
        print(f'Saved Excel: {xlsx_file}')

    except ImportError:
        print('openpyxl not available, skipping Excel output')

    # Print top 20
    print(f'\n{"=" * 80}')
    print(f'TOP 20 BUILDINGS BY HEIGHT')
    print(f'{"=" * 80}')
    print(f'{"#":>3} {"名稱":<16} {"高度":>6} {"樓層":>4} {"寬":>7} {"長":>7} {"面積":>9} {"總樓地板":>10}')
    print(f'{"-" * 80}')
    for i, r in enumerate(rows[:20], 1):
        tfa = f"{int(r['total_floor_area_m2']):,}" if r['total_floor_area_m2'] else '—'
        print(f'{i:>3} {r["name"]:<16} {r["height_m"]:>6.1f} {r["floors"]:>4} '
              f'{r["width_m"]:>6.1f}m {r["length_m"]:>6.1f}m '
              f'{r["footprint_area_m2"]:>8.1f} {tfa:>10}')


if __name__ == '__main__':
    main()
